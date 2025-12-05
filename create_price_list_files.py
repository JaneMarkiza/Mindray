#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для создания 6 файлов Excel из исходного файла Price list_actual.xlsx
с сохранением форматирования
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from copy import copy

# Путь к исходному файлу
source_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price list_actual.xlsx'

# Открываем исходный файл
print(f"Открытие файла: {source_file}")
wb_source = load_workbook(source_file, data_only=False)

print(f"Доступные страницы: {wb_source.sheetnames}")

def copy_columns_with_formatting(source_ws, dest_ws, columns_to_copy, headers, start_row=1):
    """
    Копирует указанные столбцы из исходного листа в целевой с сохранением форматирования
    
    Args:
        source_ws: исходный лист
        dest_ws: целевой лист
        columns_to_copy: список индексов столбцов для копирования (1-based)
        headers: список названий столбцов (None для использования исходного)
        start_row: строка начала данных (1 = заголовок)
    """
    # Устанавливаем заголовки
    for col_idx, (old_col_idx, header) in enumerate(zip(columns_to_copy, headers), 1):
        cell = dest_ws.cell(row=start_row, column=col_idx)
        source_header_cell = source_ws.cell(row=start_row, column=old_col_idx)
        
        # Используем указанное название или исходное из файла
        if header is not None:
            cell.value = header
        else:
            cell.value = source_header_cell.value
        
        # Копируем форматирование из первой строки исходного файла
        if source_header_cell.has_style:
            if source_header_cell.font:
                cell.font = copy(source_header_cell.font)
            if source_header_cell.fill:
                cell.fill = copy(source_header_cell.fill)
            if source_header_cell.border:
                cell.border = copy(source_header_cell.border)
            if source_header_cell.alignment:
                cell.alignment = copy(source_header_cell.alignment)
            cell.number_format = source_header_cell.number_format
    
    # Копируем данные из исходного файла (начиная со следующей строки после заголовка)
    for row_idx in range(start_row + 1, source_ws.max_row + 1):
        for new_col_idx, old_col_idx in enumerate(columns_to_copy, 1):
            source_cell = source_ws.cell(row=row_idx, column=old_col_idx)
            dest_cell = dest_ws.cell(row=row_idx, column=new_col_idx)
            
            # Копируем значение
            dest_cell.value = source_cell.value
            
            # Копируем форматирование
            if source_cell.has_style:
                if source_cell.font:
                    dest_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    dest_cell.fill = copy(source_cell.fill)
                if source_cell.border:
                    dest_cell.border = copy(source_cell.border)
                if source_cell.alignment:
                    dest_cell.alignment = copy(source_cell.alignment)
                dest_cell.number_format = source_cell.number_format
            
            # Устанавливаем перенос слов для всех ячеек
            if dest_cell.alignment is None:
                dest_cell.alignment = Alignment(wrap_text=True)
            else:
                dest_cell.alignment = Alignment(
                    horizontal=dest_cell.alignment.horizontal,
                    vertical=dest_cell.alignment.vertical,
                    wrap_text=True,
                    shrink_to_fit=dest_cell.alignment.shrink_to_fit,
                    indent=dest_cell.alignment.indent
                )
    
    # Копируем ширину столбцов из исходного файла
    for new_col_idx, old_col_idx in enumerate(columns_to_copy, 1):
        old_col_letter = get_column_letter(old_col_idx)
        new_col_letter = get_column_letter(new_col_idx)
        
        # Копируем ширину столбца
        if old_col_letter in source_ws.column_dimensions:
            dest_ws.column_dimensions[new_col_letter].width = source_ws.column_dimensions[old_col_letter].width
        elif source_ws.column_dimensions[old_col_letter].width is not None:
            dest_ws.column_dimensions[new_col_letter].width = source_ws.column_dimensions[old_col_letter].width
    
    # Копируем высоту строк
    for row_idx in range(start_row, source_ws.max_row + 1):
        if row_idx in source_ws.row_dimensions:
            dest_ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height

# ========== ФАЙЛ 1: Analyzer_1 ==========
print("\n=== Создание файла Analyzer_1 ===")
if 'Analyzers' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Analyzers' не найдена.")
else:
    ws_source = wb_source['Analyzers']
    wb_dest = Workbook()
    ws_dest = wb_dest.active
    ws_dest.title = "Analyzer_1"
    
    columns_to_copy = [1, 2, 3, 6, 7, 8]  # A, B, C, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    
    output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Analyzer_1.xlsx'
    wb_dest.save(output_file)
    print(f"Файл создан: {output_file}")

# ========== ФАЙЛ 2: Analyzer_2 ==========
print("\n=== Создание файла Analyzer_2 ===")
if 'Analyzers' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Analyzers' не найдена.")
else:
    ws_source = wb_source['Analyzers']
    wb_dest = Workbook()
    ws_dest = wb_dest.active
    ws_dest.title = "Analyzer_2"
    
    columns_to_copy = [1, 2, 4, 6, 7, 8]  # A, B, D, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    
    output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Analyzer_2.xlsx'
    wb_dest.save(output_file)
    print(f"Файл создан: {output_file}")

# ========== ФАЙЛ 3: Analyzer_3 ==========
print("\n=== Создание файла Analyzer_3 ===")
if 'Analyzers' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Analyzers' не найдена.")
else:
    ws_source = wb_source['Analyzers']
    wb_dest = Workbook()
    ws_dest = wb_dest.active
    ws_dest.title = "Analyzer_3"
    
    columns_to_copy = [1, 2, 5, 6, 7, 8]  # A, B, E, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    
    output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Analyzer_3.xlsx'
    wb_dest.save(output_file)
    print(f"Файл создан: {output_file}")

# ========== ФАЙЛ 4: Reagent_1 ==========
print("\n=== Создание файла Reagent_1 ===")
wb_dest = Workbook()
wb_dest.remove(wb_dest.active)  # Удаляем пустой лист

# Страница 1: Hematology reagents
if 'Hematology reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Hematology reagents' не найдена.")
else:
    ws_source = wb_source['Hematology reagents']
    ws_dest = wb_dest.create_sheet("Hematology reagents")
    
    columns_to_copy = [1, 2, 3, 6, 7, 8]  # A, B, C, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 1 создана: Hematology reagents")

# Страница 2: CC reagents
if 'CC reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CC reagents' не найдена.")
else:
    ws_source = wb_source['CC reagents']
    ws_dest = wb_dest.create_sheet("CC reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 7, 10, 11, 12, 13, 14]  # A, B, C, D, E, F, G, J, K, L, M, N
    headers = [
        'Group',
        'Каталожный номер',
        'Номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор',
        'Модель анализатора'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 2 создана: CC reagents")

# Страница 3: CLIA reagents
if 'CLIA reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CLIA reagents' не найдена.")
else:
    ws_source = wb_source['CLIA reagents']
    ws_dest = wb_dest.create_sheet("CLIA reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 7, 10, 11, 12, 13]  # A, B, C, D, E, F, G, J, K, L, M
    headers = [
        'Панель',
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Максимальное количество тестов из набора',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 3 создана: CLIA reagents")

# Страница 4: Coag reagents
coag_sheet_name = 'Coag reagents' if 'Coag reagents' in wb_source.sheetnames else 'Coag reagent'
if coag_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Coag reagents' или 'Coag reagent' не найдена.")
else:
    ws_source = wb_source[coag_sheet_name]
    ws_dest = wb_dest.create_sheet("Coag reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 8, 9, 10]  # A, B, C, D, E, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 4 создана: Coag reagents")

# Страница 5: Urine reagents
urine_sheet_name = 'Urine reagents' if 'Urine reagents' in wb_source.sheetnames else 'Urine reagent'
if urine_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Urine reagents' или 'Urine reagent' не найдена.")
else:
    ws_source = wb_source[urine_sheet_name]
    ws_dest = wb_dest.create_sheet("Urine reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 8, 9, 10]  # A, B, C, D, E, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 5 создана: Urine reagents")

output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Reagent_1.xlsx'
wb_dest.save(output_file)
print(f"Файл создан: {output_file}")

# ========== ФАЙЛ 5: Reagent_2 ==========
print("\n=== Создание файла Reagent_2 ===")
wb_dest = Workbook()
wb_dest.remove(wb_dest.active)  # Удаляем пустой лист

# Страница 1: Hematology reagents
if 'Hematology reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Hematology reagents' не найдена.")
else:
    ws_source = wb_source['Hematology reagents']
    ws_dest = wb_dest.create_sheet("Hematology reagents")
    
    columns_to_copy = [1, 2, 4, 6, 7, 8]  # A, B, D, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 1 создана: Hematology reagents")

# Страница 2: CC reagents
if 'CC reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CC reagents' не найдена.")
else:
    ws_source = wb_source['CC reagents']
    ws_dest = wb_dest.create_sheet("CC reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 13, 14]  # A, B, C, D, E, F, H, J, K, L, M, N
    headers = [
        'Group',
        'Каталожный номер',
        'Номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор',
        'Модель анализатора'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 2 создана: CC reagents")

# Страница 3: CLIA reagents
if 'CLIA reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CLIA reagents' не найдена.")
else:
    ws_source = wb_source['CLIA reagents']
    ws_dest = wb_dest.create_sheet("CLIA reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 13]  # A, B, C, D, E, F, H, J, K, L, M
    headers = [
        'Панель',
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Максимальное количество тестов из набора',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 3 создана: CLIA reagents")

# Страница 4: Coag reagents
coag_sheet_name = 'Coag reagents' if 'Coag reagents' in wb_source.sheetnames else 'Coag reagent'
if coag_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Coag reagents' или 'Coag reagent' не найдена.")
else:
    ws_source = wb_source[coag_sheet_name]
    ws_dest = wb_dest.create_sheet("Coag reagents")
    
    columns_to_copy = [1, 2, 3, 4, 6, 8, 9, 10]  # A, B, C, D, F, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 4 создана: Coag reagents")

# Страница 5: Urine reagents
urine_sheet_name = 'Urine reagents' if 'Urine reagents' in wb_source.sheetnames else 'Urine reagent'
if urine_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Urine reagents' или 'Urine reagent' не найдена.")
else:
    ws_source = wb_source[urine_sheet_name]
    ws_dest = wb_dest.create_sheet("Urine reagents")
    
    columns_to_copy = [1, 2, 3, 4, 6, 8, 9, 10]  # A, B, C, D, F, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 5 создана: Urine reagents")

output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Reagent_2.xlsx'
wb_dest.save(output_file)
print(f"Файл создан: {output_file}")

# ========== ФАЙЛ 6: Reagent_3 ==========
print("\n=== Создание файла Reagent_3 ===")
wb_dest = Workbook()
wb_dest.remove(wb_dest.active)  # Удаляем пустой лист

# Страница 1: Hematology reagents
if 'Hematology reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Hematology reagents' не найдена.")
else:
    ws_source = wb_source['Hematology reagents']
    ws_dest = wb_dest.create_sheet("Hematology reagents")
    
    columns_to_copy = [1, 2, 5, 6, 7, 8]  # A, B, E, F, G, H
    headers = [
        'PA number',
        'Наименование',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС',
        'Комментарии'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 1 создана: Hematology reagents")

# Страница 2: CC reagents
if 'CC reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CC reagents' не найдена.")
else:
    ws_source = wb_source['CC reagents']
    ws_dest = wb_dest.create_sheet("CC reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 14]  # A, B, C, D, E, F, I, J, K, L, M, N
    headers = [
        'Group',
        'Каталожный номер',
        'Номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор',
        'Модель анализатора'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 2 создана: CC reagents")

# Страница 3: CLIA reagents
if 'CLIA reagents' not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'CLIA reagents' не найдена.")
else:
    ws_source = wb_source['CLIA reagents']
    ws_dest = wb_dest.create_sheet("CLIA reagents")
    
    columns_to_copy = [1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13]  # A, B, C, D, E, F, I, J, K, L, M
    headers = [
        'Панель',
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Максимальное количество тестов из набора',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Контроль',
        'Калибратор'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 3 создана: CLIA reagents")

# Страница 4: Coag reagents
coag_sheet_name = 'Coag reagents' if 'Coag reagents' in wb_source.sheetnames else 'Coag reagent'
if coag_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Coag reagents' или 'Coag reagent' не найдена.")
else:
    ws_source = wb_source[coag_sheet_name]
    ws_dest = wb_dest.create_sheet("Coag reagents")
    
    columns_to_copy = [1, 2, 3, 4, 7, 8, 9, 10]  # A, B, C, D, G, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 4 создана: Coag reagents")

# Страница 5: Urine reagents
urine_sheet_name = 'Urine reagents' if 'Urine reagents' in wb_source.sheetnames else 'Urine reagent'
if urine_sheet_name not in wb_source.sheetnames:
    print(f"Ошибка: Страница 'Urine reagents' или 'Urine reagent' не найдена.")
else:
    ws_source = wb_source[urine_sheet_name]
    ws_dest = wb_dest.create_sheet("Urine reagents")
    
    columns_to_copy = [1, 2, 3, 4, 7, 8, 9, 10]  # A, B, C, D, G, H, I, J
    headers = [
        'Каталожный номер',
        'Наименование на английском языке',
        'Наименование на русском языке',
        'Фасовка',
        'Стоимость за единицу, включая НДС',
        'Рекомендованная стоимость для пользователя, включая НДС',
        'НДС%',
        'Комментарий'
    ]
    
    copy_columns_with_formatting(ws_source, ws_dest, columns_to_copy, headers)
    print("  Страница 5 создана: Urine reagents")

output_file = '/Users/markizajob/Library/Mobile Documents/com~apple~CloudDocs/Mindray 2025/Prices/Price lists emailing/Updated/Reagent_3.xlsx'
wb_dest.save(output_file)
print(f"Файл создан: {output_file}")

print("\n=== Все файлы успешно созданы! ===")

