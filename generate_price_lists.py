from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string


SOURCE_FILE = Path("/Users/evgeniamarkizova/Price list_actual.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent


def to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(" ", "").replace(",", ".")
        if not normalized:
            return None
        if normalized.endswith("%"):
            normalized = normalized[:-1]
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def parse_vat(value) -> float | None:
    number = to_float(value)
    if number is None:
        return None
    if number > 1:
        return number / 100.0
    return number


def copy_cell_style(src: Cell, dst: Cell) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def set_wrap_text(cell: Cell) -> None:
    if cell.alignment is None:
        cell.alignment = Alignment(wrap_text=True)
        return
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    cell.alignment = alignment


def copy_standard_sheet(
    src_ws_values,
    src_ws_style,
    dst_ws,
    source_columns: Iterable[str],
    headers: list[str],
) -> None:
    source_columns = list(source_columns)
    dst_ws.freeze_panes = src_ws_style.freeze_panes
    dst_ws.row_dimensions[1].height = src_ws_style.row_dimensions[1].height

    for out_col_idx, (src_col, header) in enumerate(zip(source_columns, headers), start=1):
        src_col_idx = column_index_from_string(src_col)
        dst_ws.column_dimensions[dst_ws.cell(row=1, column=out_col_idx).column_letter].width = (
            src_ws_style.column_dimensions[src_col].width
        )

        src_header_style = src_ws_style.cell(row=1, column=src_col_idx)
        dst_header = dst_ws.cell(row=1, column=out_col_idx, value=header)
        copy_cell_style(src_header_style, dst_header)
        set_wrap_text(dst_header)

        for row in range(2, src_ws_values.max_row + 1):
            src_value_cell = src_ws_values.cell(row=row, column=src_col_idx)
            src_style_cell = src_ws_style.cell(row=row, column=src_col_idx)

            dst_cell = dst_ws.cell(row=row, column=out_col_idx, value=src_value_cell.value)
            copy_cell_style(src_style_cell, dst_cell)
            set_wrap_text(dst_cell)

    for row in range(2, src_ws_values.max_row + 1):
        dst_ws.row_dimensions[row].height = src_ws_style.row_dimensions[row].height


def copy_price_wo_vat_sheet(
    src_ws_values,
    src_ws_style,
    dst_ws,
    id_col: str,
    name_col: str,
    price_cols: tuple[str, str, str],
    vat_col: str,
    headers: list[str],
) -> None:
    dst_ws.freeze_panes = src_ws_style.freeze_panes
    dst_ws.row_dimensions[1].height = src_ws_style.row_dimensions[1].height

    mapped_cols = [id_col, name_col, *price_cols]
    for out_col_idx, src_col in enumerate(mapped_cols, start=1):
        src_col_idx = column_index_from_string(src_col)
        dst_col_letter = dst_ws.cell(row=1, column=out_col_idx).column_letter
        dst_ws.column_dimensions[dst_col_letter].width = src_ws_style.column_dimensions[src_col].width

        src_header_style = src_ws_style.cell(row=1, column=src_col_idx)
        dst_header = dst_ws.cell(row=1, column=out_col_idx, value=headers[out_col_idx - 1])
        copy_cell_style(src_header_style, dst_header)
        set_wrap_text(dst_header)

    id_idx = column_index_from_string(id_col)
    name_idx = column_index_from_string(name_col)
    price_idx = [column_index_from_string(col) for col in price_cols]
    vat_idx = column_index_from_string(vat_col)

    for row in range(2, src_ws_values.max_row + 1):
        dst_ws.row_dimensions[row].height = src_ws_style.row_dimensions[row].height

        src_id_value = src_ws_values.cell(row=row, column=id_idx).value
        src_name_value = src_ws_values.cell(row=row, column=name_idx).value
        vat = parse_vat(src_ws_values.cell(row=row, column=vat_idx).value)

        id_dst = dst_ws.cell(row=row, column=1, value=src_id_value)
        name_dst = dst_ws.cell(row=row, column=2, value=src_name_value)
        copy_cell_style(src_ws_style.cell(row=row, column=id_idx), id_dst)
        copy_cell_style(src_ws_style.cell(row=row, column=name_idx), name_dst)
        set_wrap_text(id_dst)
        set_wrap_text(name_dst)

        for i, src_price_col_idx in enumerate(price_idx, start=3):
            raw_price = to_float(src_ws_values.cell(row=row, column=src_price_col_idx).value)
            price_wo_vat = None
            if raw_price is not None and vat is not None:
                denominator = 1 + vat
                if denominator != 0:
                    price_wo_vat = raw_price / denominator

            price_dst = dst_ws.cell(row=row, column=i, value=price_wo_vat)
            copy_cell_style(src_ws_style.cell(row=row, column=src_price_col_idx), price_dst)
            set_wrap_text(price_dst)


def create_output_workbook(
    wb_values,
    wb_style,
    file_name: str,
    sheet_specs: list[dict],
) -> None:
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for spec in sheet_specs:
        src_name = spec["source_sheet"]
        dst_name = spec.get("target_sheet", src_name)
        dst_ws = out_wb.create_sheet(title=dst_name[:31])

        src_ws_values = wb_values[src_name]
        src_ws_style = wb_style[src_name]

        if spec["kind"] == "standard":
            copy_standard_sheet(
                src_ws_values=src_ws_values,
                src_ws_style=src_ws_style,
                dst_ws=dst_ws,
                source_columns=spec["columns"],
                headers=spec["headers"],
            )
        elif spec["kind"] == "price_wo_vat":
            copy_price_wo_vat_sheet(
                src_ws_values=src_ws_values,
                src_ws_style=src_ws_style,
                dst_ws=dst_ws,
                id_col=spec["id_col"],
                name_col=spec["name_col"],
                price_cols=spec["price_cols"],
                vat_col=spec["vat_col"],
                headers=spec["headers"],
            )
        else:
            raise ValueError(f"Unknown spec kind: {spec['kind']}")

    out_path = OUTPUT_DIR / f"{file_name}.xlsx"
    out_wb.save(out_path)
    print(f"Created: {out_path}")


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE_FILE}")

    wb_style = load_workbook(SOURCE_FILE, data_only=False)
    wb_values = load_workbook(SOURCE_FILE, data_only=True)

    analyzer_headers = [
        "PA number",
        "Наименование",
        "Стоимость за единицу, включая НДС",
        "Рекомендованная стоимость для пользователя, включая НДС",
        "НДС",
        "Комментарии",
    ]

    reagent_h_headers = analyzer_headers
    cc_headers = [
        "Group",
        "Каталожный номер",
        "Номер",
        "Наименование на английском языке",
        "Наименование на русском языке",
        "Фасовка",
        "Стоимость за единицу, включая НДС",
        "Рекомендованная стоимость для пользователя, включая НДС",
        "НДС%",
        "Контроль",
        "Калибратор",
        "Модель анализатора",
    ]
    clia_headers = [
        "Панель",
        "Каталожный номер",
        "Наименование на английском языке",
        "Наименование на русском языке",
        "Фасовка",
        "Максимальное количество тестов из набора",
        "Стоимость за единицу, включая НДС",
        "Рекомендованная стоимость для пользователя, включая НДС",
        "НДС%",
        "Контроль",
        "Калибратор",
    ]
    coag_headers = [
        "Каталожный номер",
        "Наименование на английском языке",
        "Наименование на русском языке",
        "Фасовка",
        "Стоимость за единицу, включая НДС",
        "Рекомендованная стоимость для пользователя, включая НДС",
        "НДС%",
        "Комментарий",
    ]

    create_output_workbook(
        wb_values,
        wb_style,
        "Analyzer_1",
        [
            {
                "kind": "standard",
                "source_sheet": "Analyzers",
                "columns": ["A", "B", "C", "F", "G", "H"],
                "headers": analyzer_headers,
            }
        ],
    )
    create_output_workbook(
        wb_values,
        wb_style,
        "Analyzer_2",
        [
            {
                "kind": "standard",
                "source_sheet": "Analyzers",
                "columns": ["A", "B", "D", "F", "G", "H"],
                "headers": analyzer_headers,
            }
        ],
    )
    create_output_workbook(
        wb_values,
        wb_style,
        "Analyzer_3",
        [
            {
                "kind": "standard",
                "source_sheet": "Analyzers",
                "columns": ["A", "B", "E", "F", "G", "H"],
                "headers": analyzer_headers,
            }
        ],
    )

    create_output_workbook(
        wb_values,
        wb_style,
        "Reagent_1",
        [
            {
                "kind": "standard",
                "source_sheet": "Hematology reagents",
                "columns": ["A", "B", "C", "F", "G", "H"],
                "headers": reagent_h_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CC reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "G", "J", "K", "L", "M", "N"],
                "headers": cc_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CLIA reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "G", "J", "K", "L", "M"],
                "headers": clia_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Coag reagents",
                "columns": ["A", "B", "C", "D", "E", "H", "I", "J"],
                "headers": coag_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Urine reagents",
                "columns": ["A", "B", "C", "D", "E", "H", "I", "J"],
                "headers": coag_headers,
            },
        ],
    )

    create_output_workbook(
        wb_values,
        wb_style,
        "Reagent_2",
        [
            {
                "kind": "standard",
                "source_sheet": "Hematology reagents",
                "columns": ["A", "B", "D", "F", "G", "H"],
                "headers": reagent_h_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CC reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "H", "J", "K", "L", "M", "N"],
                "headers": cc_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CLIA reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "H", "J", "K", "L", "M"],
                "headers": clia_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Coag reagents",
                "columns": ["A", "B", "C", "D", "F", "H", "I", "J"],
                "headers": coag_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Urine reagents",
                "columns": ["A", "B", "C", "D", "F", "H", "I", "J"],
                "headers": coag_headers,
            },
        ],
    )

    create_output_workbook(
        wb_values,
        wb_style,
        "Reagent_3",
        [
            {
                "kind": "standard",
                "source_sheet": "Hematology reagents",
                "columns": ["A", "B", "E", "F", "G", "H"],
                "headers": reagent_h_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CC reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "I", "J", "K", "L", "M", "N"],
                "headers": cc_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "CLIA reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "I", "J", "K", "L", "M"],
                "headers": clia_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Coag reagents",
                "columns": ["A", "B", "C", "D", "G", "H", "I", "J"],
                "headers": coag_headers,
            },
            {
                "kind": "standard",
                "source_sheet": "Urine reagents",
                "columns": ["A", "B", "C", "D", "G", "H", "I", "J"],
                "headers": coag_headers,
            },
        ],
    )

    create_output_workbook(
        wb_values,
        wb_style,
        "Price_wo_VAT",
        [
            {
                "kind": "price_wo_vat",
                "source_sheet": "Analyzers",
                "id_col": "A",
                "name_col": "B",
                "price_cols": ("C", "D", "E"),
                "vat_col": "G",
                "headers": [
                    "PA number",
                    "Наименование",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
            {
                "kind": "price_wo_vat",
                "source_sheet": "Hematology reagents",
                "id_col": "A",
                "name_col": "B",
                "price_cols": ("C", "D", "E"),
                "vat_col": "G",
                "headers": [
                    "Наименование",
                    "Каталожный номер",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
            {
                "kind": "price_wo_vat",
                "source_sheet": "CC reagents",
                "id_col": "B",
                "name_col": "D",
                "price_cols": ("G", "H", "I"),
                "vat_col": "K",
                "headers": [
                    "Каталожный номер",
                    "Наименование",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
            {
                "kind": "price_wo_vat",
                "source_sheet": "CLIA reagents",
                "id_col": "B",
                "name_col": "C",
                "price_cols": ("G", "H", "I"),
                "vat_col": "K",
                "headers": [
                    "Каталожный номер",
                    "Наименование",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
            {
                "kind": "price_wo_vat",
                "source_sheet": "Coag reagents",
                "id_col": "A",
                "name_col": "B",
                "price_cols": ("E", "F", "G"),
                "vat_col": "I",
                "headers": [
                    "Каталожный номер",
                    "Наименование",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
            {
                "kind": "price_wo_vat",
                "source_sheet": "Urine reagents",
                "id_col": "A",
                "name_col": "B",
                "price_cols": ("E", "F", "G"),
                "vat_col": "I",
                "headers": [
                    "Каталожный номер",
                    "Наименование",
                    "A цена без НДС",
                    "B цена без НДС",
                    "C цена без НДС",
                ],
            },
        ],
    )

    create_output_workbook(
        wb_values,
        wb_style,
        "RZD",
        [
            {
                "kind": "standard",
                "source_sheet": "Hematology reagents",
                "columns": ["A", "B", "G", "H"],
                "headers": ["PA number", "Наименование", "НДС", "Комментарии"],
            },
            {
                "kind": "standard",
                "source_sheet": "CC reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "K", "L", "M", "N"],
                "headers": [
                    "Group",
                    "Каталожный номер",
                    "Номер",
                    "Наименование на английском языке",
                    "Наименование на русском языке",
                    "Фасовка",
                    "НДС%",
                    "Контроль",
                    "Калибратор",
                    "Модель анализатора",
                ],
            },
            {
                "kind": "standard",
                "source_sheet": "CLIA reagents",
                "columns": ["A", "B", "C", "D", "E", "F", "K", "L", "M"],
                "headers": [
                    "Панель",
                    "Каталожный номер",
                    "Наименование на английском языке",
                    "Наименование на русском языке",
                    "Фасовка",
                    "Максимальное количество тестов из набора",
                    "НДС%",
                    "Контроль",
                    "Калибратор",
                ],
            },
            {
                "kind": "standard",
                "source_sheet": "Coag reagents",
                "columns": ["A", "B", "C", "D", "I", "J"],
                "headers": [
                    "Каталожный номер",
                    "Наименование на английском языке",
                    "Наименование на русском языке",
                    "Фасовка",
                    "НДС%",
                    "Комментарий",
                ],
            },
            {
                "kind": "standard",
                "source_sheet": "Urine reagents",
                "columns": ["A", "B", "C", "D", "I", "J"],
                "headers": [
                    "Каталожный номер",
                    "Наименование на английском языке",
                    "Наименование на русском языке",
                    "Фасовка",
                    "НДС%",
                    "Комментарий",
                ],
            },
        ],
    )


if __name__ == "__main__":
    main()
